"""高亮标记器：将匹配到的行标记颜色高亮。"""

from copy import copy

from openpyxl.styles import PatternFill


def highlight_rows(ws, matched_rows: list, color: str = "FFFF00") -> None:
    """
    将指定行的所有单元格背景色设置为指定颜色。

    Args:
        ws: openpyxl工作表对象
        matched_rows: 需要高亮的行号列表（1-based）
        color: 十六进制颜色值，默认黄色
    """
    if not matched_rows:
        return

    fill = PatternFill(fill_type="solid", fgColor=color)
    matched_set = set(matched_rows)
    max_col = ws.max_column

    for row_num in matched_set:
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = copy(cell.font)
            cell.border = copy(cell.border)
            cell.alignment = copy(cell.alignment)
            cell.number_format = cell.number_format
            cell.protection = copy(cell.protection)
            cell.fill = fill
