"""Excel 导入器：读取 Excel 文件数据和从 Excel/文本文件导入剧名列表。"""

import os


class ExcelImporter:
    @staticmethod
    def import_file(file_path: str) -> tuple[list[str], list[list]]:
        """
        读取 Excel 文件，返回 (headers, rows)。
        headers: 第一行作为表头列名列表
        rows: 其余行数据列表的列表
        支持 .xlsx 和 .xls 格式。
        """
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"文件未找到: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".xlsx":
            return ExcelImporter._import_file_xlsx(file_path)
        elif ext == ".xls":
            return ExcelImporter._import_file_xls(file_path)
        else:
            raise ValueError(f"不支持的文件格式: '{ext}'。支持 .xlsx 和 .xls 格式")

    @staticmethod
    def import_drama_names(file_path: str, column_id: str = None) -> list[str]:
        """
        从 Excel 或文本文件导入剧名列表。
        文本文件（.txt）按行读取并去除首尾空格。
        Excel 文件读取指定列（column_id）或第一列。
        返回非空字符串列表。
        """
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"文件未找到: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".txt":
            return ExcelImporter._import_names_from_text(file_path)
        elif ext in (".xlsx", ".xls"):
            return ExcelImporter._import_names_from_excel(file_path, ext, column_id)
        else:
            raise ValueError(f"不支持的文件格式: '{ext}'。支持 .xlsx、.xls 和 .txt 格式")

    # --- private helpers ---

    @staticmethod
    def _import_file_xlsx(file_path: str) -> tuple[list[str], list[list]]:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=False)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows()

            # 读取表头
            first_row = next(rows_iter, None)
            if first_row is None:
                return [], []
            headers = [str(cell.value) if cell.value is not None else "" for cell in first_row]

            # 读取数据行
            rows = []
            for row in rows_iter:
                row_data = [ExcelImporter._convert_cell_value(cell.value) for cell in row]
                rows.append(row_data)
            return headers, rows
        finally:
            wb.close()

    @staticmethod
    def _import_file_xls(file_path: str) -> tuple[list[str], list[list]]:
        import xlrd

        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)

        if ws.nrows == 0:
            return [], []

        # 读取表头
        headers = [
            str(ws.cell_value(0, col)) if ws.cell_value(0, col) != "" else ""
            for col in range(ws.ncols)
        ]

        # 读取数据行
        rows = []
        for row_idx in range(1, ws.nrows):
            row_data = [
                ExcelImporter._convert_xls_cell(ws, row_idx, col_idx)
                for col_idx in range(ws.ncols)
            ]
            rows.append(row_data)
        return headers, rows

    @staticmethod
    def _import_names_from_text(file_path: str) -> list[str]:
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
        return [line.strip() for line in lines if line.strip()]

    @staticmethod
    def _import_names_from_excel(file_path: str, ext: str, column_id: str = None) -> list[str]:
        if ext == ".xlsx":
            return ExcelImporter._import_names_xlsx(file_path, column_id)
        else:
            return ExcelImporter._import_names_xls(file_path, column_id)

    @staticmethod
    def _import_names_xlsx(file_path: str, column_id: str = None) -> list[str]:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=False)
        try:
            ws = wb.active
            if ws.max_row is None or ws.max_row < 1:
                return []

            # 确定列索引（1-based）
            col_idx = ExcelImporter._resolve_xlsx_column(ws, column_id)

            names = []
            for row_num in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row_num, column=col_idx).value
                if cell_value is not None:
                    name = str(cell_value).strip()
                    if name:
                        names.append(name)
            return names
        finally:
            wb.close()

    @staticmethod
    def _import_names_xls(file_path: str, column_id: str = None) -> list[str]:
        import xlrd

        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)

        if ws.nrows < 1:
            return []

        # 确定列索引（0-based）
        col_idx = ExcelImporter._resolve_xls_column(ws, column_id)

        names = []
        for row_num in range(1, ws.nrows):
            cell_value = ws.cell_value(row_num, col_idx)
            if cell_value is not None:
                name = str(cell_value).strip()
                if name:
                    names.append(name)
        return names

    @staticmethod
    def _resolve_xlsx_column(ws, column_id: str = None) -> int:
        """解析 xlsx 列索引（返回 1-based）。column_id 为 None 时返回第一列。"""
        if column_id is None:
            return 1

        # 先尝试按表头名匹配
        for cell in ws[1]:
            if cell.value is not None and str(cell.value).strip() == column_id.strip():
                return cell.column

        # 再尝试解析为数字列号
        try:
            col_num = int(column_id)
            if 1 <= col_num <= ws.max_column:
                return col_num
            raise ValueError(f"列号 {col_num} 超出范围（1-{ws.max_column}）")
        except ValueError as e:
            if "超出范围" in str(e):
                raise
            headers = [str(c.value).strip() for c in ws[1] if c.value is not None]
            raise ValueError(f"列标识 '{column_id}' 未找到。可用的列名: {headers}")

    @staticmethod
    def _resolve_xls_column(ws, column_id: str = None) -> int:
        """解析 xls 列索引（返回 0-based）。column_id 为 None 时返回第一列。"""
        if column_id is None:
            return 0

        # 先尝试按表头名匹配
        for col in range(ws.ncols):
            value = ws.cell_value(0, col)
            if value is not None and str(value).strip() == column_id.strip():
                return col

        # 再尝试解析为数字列号
        try:
            col_num = int(column_id)
            if 1 <= col_num <= ws.ncols:
                return col_num - 1
            raise ValueError(f"列号 {col_num} 超出范围（1-{ws.ncols}）")
        except ValueError as e:
            if "超出范围" in str(e):
                raise
            headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols) if ws.cell_value(0, c)]
            raise ValueError(f"列标识 '{column_id}' 未找到。可用的列名: {headers}")

    @staticmethod
    def _convert_cell_value(value):
        """将 openpyxl 单元格值转换为合适的 Python 类型。"""
        if value is None:
            return None
        return value

    @staticmethod
    def _convert_xls_cell(ws, row_idx, col_idx):
        """将 xlrd 单元格值转换为合适的 Python 类型。"""
        import xlrd

        cell_type = ws.cell_type(row_idx, col_idx)
        value = ws.cell_value(row_idx, col_idx)

        if cell_type == xlrd.XL_CELL_EMPTY:
            return None
        elif cell_type == xlrd.XL_CELL_NUMBER:
            # xlrd 将所有数字存为 float，如果是整数则转为 int
            if value == int(value):
                return int(value)
            return value
        elif cell_type == xlrd.XL_CELL_TEXT:
            return value
        elif cell_type == xlrd.XL_CELL_BOOLEAN:
            return bool(value)
        else:
            return value
