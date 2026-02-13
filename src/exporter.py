"""数据导出器：将数据导出为 Excel 文件。"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill


YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


class Exporter:
    @staticmethod
    def export_to_excel(file_path: str, headers: list[str], rows: list[list]) -> None:
        """将数据导出为 Excel 文件。"""
        wb = Workbook()
        ws = wb.active
        if headers:
            ws.append(headers)
        for row in rows:
            ws.append(row)
        wb.save(file_path)

    @staticmethod
    def export_with_highlight(file_path: str, headers: list[str], rows: list[list],
                              matched_indices: list[int]) -> None:
        """导出完整表格，匹配行高亮黄色。"""
        wb = Workbook()
        ws = wb.active
        if headers:
            ws.append(headers)

        matched_set = set(matched_indices)
        for i, row in enumerate(rows):
            ws.append(row)
            if i in matched_set:
                excel_row = i + 2  # +1 表头 +1 openpyxl 1-based
                for col in range(1, len(headers) + 1):
                    ws.cell(row=excel_row, column=col).fill = YELLOW_FILL

        wb.save(file_path)
