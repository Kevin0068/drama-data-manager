"""Excel读取器：从Excel文件的指定列读取剧名数据。支持 .xlsx 和 .xls 格式。"""

import os
from openpyxl import load_workbook


def resolve_column_index(ws, column_id: str) -> int:
    """
    将列标识解析为列索引（1-based）。
    优先按列名匹配第一行表头，若无匹配则尝试解析为数字列号。
    """
    header_names = []
    for cell in ws[1]:
        value = cell.value
        if value is not None:
            header_names.append(str(value).strip())
            if str(value).strip() == column_id.strip():
                return cell.column
        else:
            header_names.append("")

    try:
        col_num = int(column_id)
        if 1 <= col_num <= ws.max_column:
            return col_num
        raise ValueError(f"列号 {col_num} 超出范围（1-{ws.max_column}）")
    except ValueError as e:
        if "超出范围" in str(e):
            raise
        available = [name for name in header_names if name]
        raise ValueError(f"列标识 '{column_id}' 未找到。可用的列名: {available}")


def read_drama_names(file_path: str, column_id: str, sheet_name: str = None) -> list[tuple[int, str]]:
    """
    从Excel文件的指定列读取剧名列表。支持 .xlsx 和 .xls 格式。
    """
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"文件未找到: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xls":
        return _read_xls(file_path, column_id, sheet_name)
    elif ext == ".xlsx":
        return _read_xlsx(file_path, column_id, sheet_name)
    else:
        raise ValueError(f"不支持的文件格式: '{ext}'。支持 .xlsx 和 .xls 格式")


def _read_xlsx(file_path, column_id, sheet_name=None):
    """读取 .xlsx 文件。"""
    wb = load_workbook(file_path, read_only=False)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {wb.sheetnames}")
            ws = wb[sheet_name]
        else:
            ws = wb.active

        col_idx = resolve_column_index(ws, column_id)

        results = []
        for row_num in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=col_idx).value
            if cell_value is not None:
                name = str(cell_value).strip()
                if name:
                    results.append((row_num, name))
        return results
    finally:
        wb.close()


def _read_xls(file_path, column_id, sheet_name=None):
    """读取 .xls 文件（使用 xlrd）。"""
    import xlrd

    wb = xlrd.open_workbook(file_path)
    if sheet_name:
        if sheet_name not in wb.sheet_names():
            raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {wb.sheet_names()}")
        ws = wb.sheet_by_name(sheet_name)
    else:
        ws = wb.sheet_by_index(0)

    col_idx = _resolve_column_xls(ws, column_id)

    results = []
    for row_num in range(1, ws.nrows):
        cell_value = ws.cell_value(row_num, col_idx)
        if cell_value is not None:
            name = str(cell_value).strip()
            if name:
                results.append((row_num + 1, name))  # 1-based row number
    return results


def _resolve_column_xls(ws, column_id):
    """解析 .xls 文件的列索引（0-based）。"""
    header_names = []
    for col in range(ws.ncols):
        value = ws.cell_value(0, col)
        if value is not None:
            name = str(value).strip()
            header_names.append(name)
            if name == column_id.strip():
                return col
        else:
            header_names.append("")

    try:
        col_num = int(column_id)
        if 1 <= col_num <= ws.ncols:
            return col_num - 1
        raise ValueError(f"列号 {col_num} 超出范围（1-{ws.ncols}）")
    except ValueError as e:
        if "超出范围" in str(e):
            raise
        available = [n for n in header_names if n]
        raise ValueError(f"列标识 '{column_id}' 未找到。可用的列名: {available}")
