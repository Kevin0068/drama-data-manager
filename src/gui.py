"""Excel剧名匹配工具 - 可视化界面"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import os

from openpyxl import load_workbook

from src.reader import read_drama_names
from src.matcher import match_dramas
from src.highlighter import highlight_rows
from src.writer import save_workbook


class DramaMatcherApp:
    def __init__(self, root):
        root.title("Excel剧名匹配工具")
        root.geometry("600x480")
        root.resizable(False, False)

        pad = {"padx": 10, "pady": 5}

        # A文档
        tk.Label(root, text="A文档（主表）:").grid(row=0, column=0, sticky="w", **pad)
        self.master_path = tk.StringVar()
        tk.Entry(root, textvariable=self.master_path, width=45).grid(row=0, column=1, **pad)
        tk.Button(root, text="浏览", command=lambda: self._browse(self.master_path)).grid(row=0, column=2, **pad)

        # A文档列名
        tk.Label(root, text="A文档剧名列:").grid(row=1, column=0, sticky="w", **pad)
        self.master_col = tk.StringVar(value="剧名")
        tk.Entry(root, textvariable=self.master_col, width=20).grid(row=1, column=1, sticky="w", **pad)

        # B文档
        tk.Label(root, text="B文档（查找列表）:").grid(row=2, column=0, sticky="w", **pad)
        self.lookup_path = tk.StringVar()
        tk.Entry(root, textvariable=self.lookup_path, width=45).grid(row=2, column=1, **pad)
        tk.Button(root, text="浏览", command=lambda: self._browse(self.lookup_path)).grid(row=2, column=2, **pad)

        # B文档列名
        tk.Label(root, text="B文档剧名列:").grid(row=3, column=0, sticky="w", **pad)
        self.lookup_col = tk.StringVar(value="剧名")
        tk.Entry(root, textvariable=self.lookup_col, width=20).grid(row=3, column=1, sticky="w", **pad)

        # 执行按钮
        self.run_btn = tk.Button(root, text="开始匹配", command=self._run, width=15, height=2,
                                 bg="#4CAF50", fg="white", font=("Microsoft YaHei", 11, "bold"))
        self.run_btn.grid(row=4, column=0, columnspan=3, pady=15)

        # 进度条
        self.progress = ttk.Progressbar(root, mode="indeterminate", length=500)
        self.progress.grid(row=5, column=0, columnspan=3, padx=10)

        # 结果显示
        self.result_text = tk.Text(root, height=10, width=70, state="disabled", font=("Consolas", 10))
        self.result_text.grid(row=6, column=0, columnspan=3, padx=10, pady=10)

    def _browse(self, var):
        path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if path:
            var.set(path)

    def _log(self, msg):
        self.result_text.config(state="normal")
        self.result_text.insert("end", msg + "\n")
        self.result_text.see("end")
        self.result_text.config(state="disabled")

    def _run(self):
        master = self.master_path.get().strip()
        lookup = self.lookup_path.get().strip()
        mcol = self.master_col.get().strip()
        lcol = self.lookup_col.get().strip()

        if not master or not lookup:
            messagebox.showwarning("提示", "请选择A文档和B文档")
            return
        if not mcol or not lcol:
            messagebox.showwarning("提示", "请填写剧名列标识")
            return

        self.run_btn.config(state="disabled")
        self.result_text.config(state="normal")
        self.result_text.delete("1.0", "end")
        self.result_text.config(state="disabled")
        self.progress.start()

        threading.Thread(target=self._execute, args=(master, lookup, mcol, lcol), daemon=True).start()

    def _execute(self, master, lookup, mcol, lcol):
        try:
            self._log(f"正在读取A文档: {os.path.basename(master)}")
            master_names = read_drama_names(master, mcol)

            self._log(f"正在读取B文档: {os.path.basename(lookup)}")
            lookup_tuples = read_drama_names(lookup, lcol)
            lookup_set = {name for _, name in lookup_tuples}

            lookup_total_raw = len(lookup_tuples)
            lookup_duplicates = lookup_total_raw - len(lookup_set)

            self._log("正在匹配...")
            result = match_dramas(master_names, lookup_set)

            if result.match_count > 0:
                self._log(f"正在高亮标记 {result.match_count} 行...")
                ext = os.path.splitext(master)[1].lower()
                if ext == ".xls":
                    # .xls 文件需要用 xlrd 读取后转为 openpyxl 工作簿处理
                    import xlrd
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, Alignment
                    xls_wb = xlrd.open_workbook(master, formatting_info=True)
                    xls_ws = xls_wb.sheet_by_index(0)
                    wb = Workbook()
                    ws = wb.active
                    for row in range(xls_ws.nrows):
                        for col in range(xls_ws.ncols):
                            ws.cell(row=row + 1, column=col + 1, value=xls_ws.cell_value(row, col))
                    highlight_rows(ws, result.matched_rows)
                    save_path = os.path.splitext(master)[0] + ".xlsx"
                    save_workbook(wb, save_path)
                    self._log(f"注意: .xls文件已转换为 .xlsx 格式保存")
                else:
                    wb = load_workbook(master)
                    ws = wb.active
                    highlight_rows(ws, result.matched_rows)
                    save_workbook(wb, master)

            self._log("=" * 40)
            self._log(f"A文档总数: {result.total_master}")
            self._log(f"B文档总数: {lookup_total_raw}（去重后: {result.total_lookup}，重复: {lookup_duplicates}）")
            self._log(f"匹配数量: {result.match_count}")

            if result.match_count == 0:
                self._log("未找到匹配项")
            else:
                self._log(f"已在原文件中标记完成: {os.path.basename(master)}")
                self._log(f"匹配的剧名: {', '.join(result.matched_names[:20])}")
                if len(result.matched_names) > 20:
                    self._log(f"  ...等共 {result.match_count} 个")

            messagebox.showinfo("完成", f"匹配完成，共找到 {result.match_count} 个匹配项")

        except (FileNotFoundError, ValueError) as e:
            self._log(f"错误: {e}")
            messagebox.showerror("错误", str(e))
        except Exception as e:
            self._log(f"未知错误: {e}")
            messagebox.showerror("错误", f"发生未知错误:\n{e}")
        finally:
            self.progress.stop()
            self.run_btn.config(state="normal")


def main():
    root = tk.Tk()
    DramaMatcherApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
