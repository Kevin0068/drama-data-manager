"""Tests for the Excel writer module."""

import os
import pytest
from openpyxl import Workbook, load_workbook

from src.writer import save_workbook


class TestSaveWorkbook:
    """Tests for save_workbook."""

    def test_file_created(self, tmp_dir):
        """Workbook is saved to the specified path."""
        wb = Workbook()
        ws = wb.active
        ws.append(["剧名"])
        ws.append(["琅琊榜"])
        path = os.path.join(tmp_dir, "output.xlsx")
        save_workbook(wb, path)
        assert os.path.exists(path)

    def test_data_intact_after_save(self, tmp_dir):
        """Saved file can be reopened and data is intact."""
        wb = Workbook()
        ws = wb.active
        ws.append(["剧名", "类型", "年份"])
        ws.append(["琅琊榜", "古装", 2015])
        ws.append(["甄嬛传", "古装", 2011])
        path = os.path.join(tmp_dir, "output.xlsx")
        save_workbook(wb, path)

        wb2 = load_workbook(path)
        ws2 = wb2.active
        assert ws2.cell(row=1, column=1).value == "剧名"
        assert ws2.cell(row=2, column=1).value == "琅琊榜"
        assert ws2.cell(row=3, column=1).value == "甄嬛传"
        assert ws2.cell(row=2, column=2).value == "古装"
        assert ws2.cell(row=2, column=3).value == 2015

    def test_overwrites_existing_file(self, tmp_dir):
        """Saving to an existing path overwrites the file."""
        path = os.path.join(tmp_dir, "output.xlsx")

        # Create initial file
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.append(["original"])
        save_workbook(wb1, path)

        # Overwrite with new content
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(["updated"])
        save_workbook(wb2, path)

        wb3 = load_workbook(path)
        ws3 = wb3.active
        assert ws3.cell(row=1, column=1).value == "updated"
