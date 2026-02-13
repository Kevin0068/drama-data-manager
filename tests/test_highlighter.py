"""Tests for the highlighter module."""

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill

from src.highlighter import highlight_rows


class TestHighlightRows:
    """Tests for highlight_rows."""

    def _make_ws(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["剧名", "类型", "年份"])
        ws.append(["琅琊榜", "古装", 2015])
        ws.append(["甄嬛传", "古装", 2011])
        ws.append(["人民的名义", "现代", 2017])
        return ws

    def test_basic_highlight_applies_yellow_fill(self):
        ws = self._make_ws()
        highlight_rows(ws, [2])

        yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col)
            assert cell.fill == yellow

    def test_unmatched_rows_unchanged(self):
        ws = self._make_ws()
        yellow = PatternFill(fill_type="solid", fgColor="FFFF00")

        highlight_rows(ws, [2])

        # Unmatched rows should NOT have yellow fill
        for row in [1, 3, 4]:
            for col in range(1, ws.max_column + 1):
                assert ws.cell(row=row, column=col).fill != yellow

    def test_existing_formatting_preserved(self):
        ws = self._make_ws()
        bold_font = Font(name="Arial", bold=True, size=14)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell = ws.cell(row=2, column=1)
        cell.font = bold_font
        cell.border = thin_border

        highlight_rows(ws, [2])

        cell = ws.cell(row=2, column=1)
        assert cell.font.bold is True
        assert cell.font.name == "Arial"
        assert cell.font.size == 14
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "thin"
        assert cell.fill == PatternFill(fill_type="solid", fgColor="FFFF00")

    def test_multiple_rows_highlighted(self):
        ws = self._make_ws()
        highlight_rows(ws, [2, 4])

        yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
        for row in [2, 4]:
            for col in range(1, ws.max_column + 1):
                assert ws.cell(row=row, column=col).fill == yellow

        # Row 3 should not be highlighted
        for col in range(1, ws.max_column + 1):
            assert ws.cell(row=3, column=col).fill != yellow

    def test_empty_matched_rows(self):
        ws = self._make_ws()
        yellow = PatternFill(fill_type="solid", fgColor="FFFF00")

        highlight_rows(ws, [])

        # No rows should have yellow fill
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                assert ws.cell(row=row, column=col).fill != yellow
