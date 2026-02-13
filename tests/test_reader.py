"""Tests for the Excel reader module."""

import os
import pytest
from openpyxl import Workbook

from src.reader import resolve_column_index, read_drama_names


class TestResolveColumnIndex:
    """Tests for resolve_column_index."""

    def test_match_header_name(self, sample_workbook):
        ws = sample_workbook.active
        assert resolve_column_index(ws, "剧名") == 1
        assert resolve_column_index(ws, "类型") == 2
        assert resolve_column_index(ws, "年份") == 3

    def test_match_header_name_with_whitespace(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["  剧名  ", "类型"])
        # column_id is stripped, header value is stripped
        assert resolve_column_index(ws, "剧名") == 1

    def test_numeric_column_id(self, sample_workbook):
        ws = sample_workbook.active
        assert resolve_column_index(ws, "1") == 1
        assert resolve_column_index(ws, "3") == 3

    def test_numeric_column_out_of_range(self, sample_workbook):
        ws = sample_workbook.active
        with pytest.raises(ValueError, match="超出范围"):
            resolve_column_index(ws, "99")

    def test_invalid_column_id(self, sample_workbook):
        ws = sample_workbook.active
        with pytest.raises(ValueError, match="未找到"):
            resolve_column_index(ws, "不存在的列")

    def test_header_priority_over_numeric(self):
        """Header name match takes priority over numeric parsing."""
        wb = Workbook()
        ws = wb.active
        ws.append(["1", "2", "data"])
        # "1" matches header name in column 1
        assert resolve_column_index(ws, "1") == 1


class TestReadDramaNames:
    """Tests for read_drama_names."""

    def _save_wb(self, wb, tmp_dir, name="test.xlsx"):
        path = os.path.join(tmp_dir, name)
        wb.save(path)
        return path

    def test_basic_read(self, sample_workbook, tmp_dir):
        path = self._save_wb(sample_workbook, tmp_dir)
        result = read_drama_names(path, "剧名")
        assert result == [(2, "琅琊榜"), (3, "甄嬛传"), (4, "人民的名义")]

    def test_read_by_column_number(self, sample_workbook, tmp_dir):
        path = self._save_wb(sample_workbook, tmp_dir)
        result = read_drama_names(path, "1")
        assert result == [(2, "琅琊榜"), (3, "甄嬛传"), (4, "人民的名义")]

    def test_skip_empty_cells(self, tmp_dir):
        wb = Workbook()
        ws = wb.active
        ws.append(["剧名"])
        ws.append(["琅琊榜"])
        ws.append([None])
        ws.append([""])
        ws.append(["  "])
        ws.append(["甄嬛传"])
        path = self._save_wb(wb, tmp_dir)
        result = read_drama_names(path, "剧名")
        assert result == [(2, "琅琊榜"), (6, "甄嬛传")]

    def test_strip_whitespace(self, tmp_dir):
        wb = Workbook()
        ws = wb.active
        ws.append(["剧名"])
        ws.append(["  琅琊榜  "])
        ws.append(["甄嬛传\t"])
        path = self._save_wb(wb, tmp_dir)
        result = read_drama_names(path, "剧名")
        assert result == [(2, "琅琊榜"), (3, "甄嬛传")]

    def test_file_not_found(self):
        with pytest.raises(FileNotFoundError, match="文件未找到"):
            read_drama_names("/nonexistent/path.xlsx", "剧名")

    def test_invalid_format(self, tmp_dir):
        path = os.path.join(tmp_dir, "test.csv")
        with open(path, "w") as f:
            f.write("a,b,c")
        with pytest.raises(ValueError, match="不支持的文件格式"):
            read_drama_names(path, "剧名")

    def test_column_not_found(self, sample_workbook, tmp_dir):
        path = self._save_wb(sample_workbook, tmp_dir)
        with pytest.raises(ValueError, match="未找到"):
            read_drama_names(path, "不存在")

    def test_specific_sheet(self, tmp_dir):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["剧名"])
        ws1.append(["琅琊榜"])
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["剧名"])
        ws2.append(["甄嬛传"])
        path = self._save_wb(wb, tmp_dir)
        result = read_drama_names(path, "剧名", sheet_name="Sheet2")
        assert result == [(2, "甄嬛传")]

    def test_invalid_sheet_name(self, sample_workbook, tmp_dir):
        path = self._save_wb(sample_workbook, tmp_dir)
        with pytest.raises(ValueError, match="不存在"):
            read_drama_names(path, "剧名", sheet_name="NoSheet")

    def test_empty_workbook(self, tmp_dir):
        wb = Workbook()
        ws = wb.active
        ws.append(["剧名"])
        path = self._save_wb(wb, tmp_dir)
        result = read_drama_names(path, "剧名")
        assert result == []
