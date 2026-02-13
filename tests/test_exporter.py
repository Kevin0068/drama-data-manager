"""Exporter 单元测试：验证 Excel 导出功能。"""

import os
import pytest
from openpyxl import load_workbook
from src.exporter import Exporter


@pytest.fixture
def tmp_xlsx(tmp_path):
    """返回临时 xlsx 文件路径。"""
    return str(tmp_path / "output.xlsx")


class TestExportToExcel:
    def test_basic_export_creates_valid_file(self, tmp_xlsx):
        """基本导出应创建有效的 Excel 文件。"""
        headers = ["姓名", "年龄", "城市"]
        rows = [["张三", 25, "北京"], ["李四", 30, "上海"]]

        Exporter.export_to_excel(tmp_xlsx, headers, rows)

        assert os.path.isfile(tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        assert ws.max_row == 3  # 1 header + 2 data rows
        assert ws.max_column == 3
        wb.close()

    def test_headers_and_data_preserved(self, tmp_xlsx):
        """导出后表头和数据应完整保留。"""
        headers = ["合集名称", "播放量", "评分"]
        rows = [["剧名A", 1000, 9.5], ["剧名B", 2000, 8.0]]

        Exporter.export_to_excel(tmp_xlsx, headers, rows)

        wb = load_workbook(tmp_xlsx)
        ws = wb.active

        # 验证表头
        actual_headers = [cell.value for cell in ws[1]]
        assert actual_headers == headers

        # 验证数据行
        actual_row1 = [cell.value for cell in ws[2]]
        assert actual_row1 == ["剧名A", 1000, 9.5]

        actual_row2 = [cell.value for cell in ws[3]]
        assert actual_row2 == ["剧名B", 2000, 8.0]
        wb.close()

    def test_empty_data_export(self, tmp_xlsx):
        """空数据导出应创建仅含表头的文件。"""
        headers = ["列A", "列B"]
        rows = []

        Exporter.export_to_excel(tmp_xlsx, headers, rows)

        assert os.path.isfile(tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        assert ws.max_row == 1  # only header
        actual_headers = [cell.value for cell in ws[1]]
        assert actual_headers == headers
        wb.close()

    def test_empty_headers_and_rows(self, tmp_xlsx):
        """完全空的导出应创建空文件。"""
        Exporter.export_to_excel(tmp_xlsx, [], [])

        assert os.path.isfile(tmp_xlsx)
        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        # openpyxl reports max_row=1 and max_column=1 for empty sheets,
        # but the cell value should be None
        assert ws.cell(1, 1).value is None
        wb.close()

    def test_none_values_in_rows(self, tmp_xlsx):
        """包含 None 值的行应正确导出。"""
        headers = ["A", "B", "C"]
        rows = [[None, "数据", None], [1, None, 3]]

        Exporter.export_to_excel(tmp_xlsx, headers, rows)

        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        actual_row1 = [cell.value for cell in ws[2]]
        assert actual_row1 == [None, "数据", None]
        actual_row2 = [cell.value for cell in ws[3]]
        assert actual_row2 == [1, None, 3]
        wb.close()
