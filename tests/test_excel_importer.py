"""Tests for ExcelImporter."""

import os
import pytest
from openpyxl import Workbook

from src.excel_importer import ExcelImporter


class TestImportFile:
    """Tests for ExcelImporter.import_file."""

    def _save(self, wb, tmp_dir, name="test.xlsx"):
        path = os.path.join(tmp_dir, name)
        wb.save(path)
        return path

    def test_basic_import(self, sample_workbook, tmp_dir):
        path = self._save(sample_workbook, tmp_dir)
        headers, rows = ExcelImporter.import_file(path)
        assert headers == ["剧名", "类型", "年份"]
        assert len(rows) == 3
        assert rows[0] == ["琅琊榜", "古装", 2015]
        assert rows[1] == ["甄嬛传", "古装", 2011]
        assert rows[2] == ["人民的名义", "现代", 2017]

    def test_empty_workbook(self, tmp_dir):
        wb = Workbook()
        ws = wb.active
        # completely empty — no rows at all
        # openpyxl always creates at least one sheet, but no data
        path = self._save(wb, tmp_dir)
        headers, rows = ExcelImporter.import_file(path)
        # An empty workbook with no data written has max_row=None or 1 empty row
        # iter_rows on empty sheet yields nothing
        assert headers == []
        assert rows == []

    def test_header_only(self, tmp_dir):
        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B", "C"])
        path = self._save(wb, tmp_dir)
        headers, rows = ExcelImporter.import_file(path)
        assert headers == ["A", "B", "C"]
        assert rows == []

    def test_none_cells(self, tmp_dir):
        wb = Workbook()
        ws = wb.active
        ws.append(["名称", "值"])
        ws.append(["test", None])
        ws.append([None, 42])
        path = self._save(wb, tmp_dir)
        headers, rows = ExcelImporter.import_file(path)
        assert headers == ["名称", "值"]
        assert rows[0] == ["test", None]
        assert rows[1] == [None, 42]

    def test_mixed_types(self, tmp_dir):
        wb = Workbook()
        ws = wb.active
        ws.append(["str_col", "int_col", "float_col", "bool_col"])
        ws.append(["hello", 42, 3.14, True])
        path = self._save(wb, tmp_dir)
        headers, rows = ExcelImporter.import_file(path)
        assert rows[0] == ["hello", 42, 3.14, True]

    def test_none_header_becomes_empty_string(self, tmp_dir):
        wb = Workbook()
        ws = wb.active
        ws.append(["A", None, "C"])
        ws.append([1, 2, 3])
        path = self._save(wb, tmp_dir)
        headers, rows = ExcelImporter.import_file(path)
        assert headers == ["A", "", "C"]

    def test_file_not_found(self):
        with pytest.raises(FileNotFoundError, match="文件未找到"):
            ExcelImporter.import_file("/nonexistent/path.xlsx")

    def test_unsupported_format(self, tmp_dir):
        path = os.path.join(tmp_dir, "test.csv")
        with open(path, "w") as f:
            f.write("a,b,c")
        with pytest.raises(ValueError, match="不支持的文件格式"):
            ExcelImporter.import_file(path)


class TestImportDramaNames:
    """Tests for ExcelImporter.import_drama_names."""

    def _save(self, wb, tmp_dir, name="test.xlsx"):
        path = os.path.join(tmp_dir, name)
        wb.save(path)
        return path

    def test_from_text_file(self, tmp_dir):
        path = os.path.join(tmp_dir, "names.txt")
        with open(path, "w", encoding="utf-8") as f:
            f.write("琅琊榜\n甄嬛传\n人民的名义\n")
        result = ExcelImporter.import_drama_names(path)
        assert result == ["琅琊榜", "甄嬛传", "人民的名义"]

    def test_text_file_strips_whitespace(self, tmp_dir):
        path = os.path.join(tmp_dir, "names.txt")
        with open(path, "w", encoding="utf-8") as f:
            f.write("  琅琊榜  \n\n  \n甄嬛传\n")
        result = ExcelImporter.import_drama_names(path)
        assert result == ["琅琊榜", "甄嬛传"]

    def test_text_file_empty(self, tmp_dir):
        path = os.path.join(tmp_dir, "empty.txt")
        with open(path, "w", encoding="utf-8") as f:
            f.write("")
        result = ExcelImporter.import_drama_names(path)
        assert result == []

    def test_from_excel_first_column(self, tmp_dir):
        wb = Workbook()
        ws = wb.active
        ws.append(["剧名", "类型"])
        ws.append(["琅琊榜", "古装"])
        ws.append(["甄嬛传", "古装"])
        path = self._save(wb, tmp_dir)
        result = ExcelImporter.import_drama_names(path)
        assert result == ["琅琊榜", "甄嬛传"]

    def test_from_excel_specified_column_by_name(self, tmp_dir):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "剧名", "类型"])
        ws.append([1, "琅琊榜", "古装"])
        ws.append([2, "甄嬛传", "古装"])
        path = self._save(wb, tmp_dir)
        result = ExcelImporter.import_drama_names(path, column_id="剧名")
        assert result == ["琅琊榜", "甄嬛传"]

    def test_from_excel_specified_column_by_number(self, tmp_dir):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "剧名"])
        ws.append([1, "琅琊榜"])
        ws.append([2, "甄嬛传"])
        path = self._save(wb, tmp_dir)
        result = ExcelImporter.import_drama_names(path, column_id="2")
        assert result == ["琅琊榜", "甄嬛传"]

    def test_from_excel_skips_empty(self, tmp_dir):
        wb = Workbook()
        ws = wb.active
        ws.append(["剧名"])
        ws.append(["琅琊榜"])
        ws.append([None])
        ws.append([""])
        ws.append(["  "])
        ws.append(["甄嬛传"])
        path = self._save(wb, tmp_dir)
        result = ExcelImporter.import_drama_names(path)
        assert result == ["琅琊榜", "甄嬛传"]

    def test_file_not_found(self):
        with pytest.raises(FileNotFoundError, match="文件未找到"):
            ExcelImporter.import_drama_names("/nonexistent/path.xlsx")

    def test_unsupported_format(self, tmp_dir):
        path = os.path.join(tmp_dir, "test.json")
        with open(path, "w") as f:
            f.write("{}")
        with pytest.raises(ValueError, match="不支持的文件格式"):
            ExcelImporter.import_drama_names(path)

    def test_invalid_column_id(self, tmp_dir):
        wb = Workbook()
        ws = wb.active
        ws.append(["剧名", "类型"])
        ws.append(["琅琊榜", "古装"])
        path = self._save(wb, tmp_dir)
        with pytest.raises(ValueError, match="未找到"):
            ExcelImporter.import_drama_names(path, column_id="不存在")
