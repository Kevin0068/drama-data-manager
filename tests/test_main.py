"""Tests for the main CLI entry point."""

import os
import tempfile

import pytest
from openpyxl import Workbook, load_workbook

from src.main import main, parse_args


@pytest.fixture
def excel_pair(tmp_path):
    """Create a master and lookup Excel file pair for testing."""
    master_path = str(tmp_path / "master.xlsx")
    lookup_path = str(tmp_path / "lookup.xlsx")

    # Master: has header + 4 drama names
    wb_m = Workbook()
    ws_m = wb_m.active
    ws_m.append(["剧名", "类型"])
    ws_m.append(["琅琊榜", "古装"])
    ws_m.append(["甄嬛传", "古装"])
    ws_m.append(["人民的名义", "现代"])
    ws_m.append(["庆余年", "古装"])
    wb_m.save(master_path)

    # Lookup: has header + 2 drama names (overlap: 琅琊榜, 人民的名义)
    wb_l = Workbook()
    ws_l = wb_l.active
    ws_l.append(["剧名", "评分"])
    ws_l.append(["琅琊榜", 9.2])
    ws_l.append(["人民的名义", 8.3])
    wb_l.save(lookup_path)

    return master_path, lookup_path


class TestParseArgs:
    def test_all_required_args(self):
        args = parse_args([
            "--master", "a.xlsx",
            "--lookup", "b.xlsx",
            "--master-col", "剧名",
            "--lookup-col", "剧名",
        ])
        assert args.master == "a.xlsx"
        assert args.lookup == "b.xlsx"
        assert args.master_col == "剧名"
        assert args.lookup_col == "剧名"

    def test_missing_required_arg_exits(self):
        with pytest.raises(SystemExit):
            parse_args(["--master", "a.xlsx"])


class TestMainFlow:
    def test_successful_match_highlights_and_saves(self, excel_pair, capsys):
        master_path, lookup_path = excel_pair
        main([
            "--master", master_path,
            "--lookup", lookup_path,
            "--master-col", "剧名",
            "--lookup-col", "剧名",
        ])
        output = capsys.readouterr().out
        assert "匹配数量: 2" in output
        assert "A文档总数: 4" in output
        assert "B文档总数: 2" in output

        # Verify highlighting was applied
        wb = load_workbook(master_path)
        ws = wb.active
        # Row 2 (琅琊榜) and Row 4 (人民的名义) should be highlighted
        assert ws.cell(row=2, column=1).fill.fgColor.rgb == "00FFFF00"
        assert ws.cell(row=4, column=1).fill.fgColor.rgb == "00FFFF00"
        # Row 3 (甄嬛传) should NOT be highlighted
        assert ws.cell(row=3, column=1).fill.fgColor.rgb != "00FFFF00"
        wb.close()

    def test_no_match_prints_message(self, tmp_path, capsys):
        master_path = str(tmp_path / "master.xlsx")
        lookup_path = str(tmp_path / "lookup.xlsx")

        wb_m = Workbook()
        ws_m = wb_m.active
        ws_m.append(["剧名"])
        ws_m.append(["琅琊榜"])
        wb_m.save(master_path)

        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l.append(["剧名"])
        ws_l.append(["三体"])
        wb_l.save(lookup_path)

        main([
            "--master", master_path,
            "--lookup", lookup_path,
            "--master-col", "剧名",
            "--lookup-col", "剧名",
        ])
        output = capsys.readouterr().out
        assert "匹配数量: 0" in output
        assert "未找到匹配项" in output

    def test_file_not_found_exits_with_code_1(self, tmp_path):
        with pytest.raises(SystemExit) as exc_info:
            main([
                "--master", str(tmp_path / "nonexistent.xlsx"),
                "--lookup", str(tmp_path / "also_missing.xlsx"),
                "--master-col", "剧名",
                "--lookup-col", "剧名",
            ])
        assert exc_info.value.code == 1

    def test_invalid_column_exits_with_code_1(self, excel_pair):
        master_path, lookup_path = excel_pair
        with pytest.raises(SystemExit) as exc_info:
            main([
                "--master", master_path,
                "--lookup", lookup_path,
                "--master-col", "不存在的列",
                "--lookup-col", "剧名",
            ])
        assert exc_info.value.code == 1

    def test_column_by_number(self, excel_pair, capsys):
        master_path, lookup_path = excel_pair
        # Column 1 is 剧名 in both files
        main([
            "--master", master_path,
            "--lookup", lookup_path,
            "--master-col", "1",
            "--lookup-col", "1",
        ])
        output = capsys.readouterr().out
        assert "匹配数量: 2" in output
